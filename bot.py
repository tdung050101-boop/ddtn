from __future__ import annotations

import argparse
import asyncio
import json
import logging
import os
import shutil
import sqlite3
import sys
import unicodedata
from dataclasses import dataclass
from datetime import date, datetime, timedelta, timezone
from pathlib import Path
from typing import Any
from zoneinfo import ZoneInfo

import discord
from discord import app_commands
from discord.ext import tasks
from dotenv import load_dotenv
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from attendance_core import (
    Shift,
    attendance_column,
    extract_stt,
    format_duration,
    manual_required_seconds,
    parse_hhmm,
    shift_window,
)


BASE_DIR = Path(__file__).resolve().parent
CONFIG_PATH = BASE_DIR / "config.json"
TEMPLATE_WORKBOOK_PATH = BASE_DIR / "template" / "Tuyet_Nguyet.xlsx"
UTC = timezone.utc

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)s | %(name)s | %(message)s",
    handlers=[logging.StreamHandler()],
)
log = logging.getLogger("tuyet-nguyet-attendance")


@dataclass(frozen=True)
class SpecialActivity:
    key: str
    label: str
    weekday: int
    column: int

    def as_shift(self) -> Shift:
        # Ca chỉ mở bằng lệnh; giờ giả không được scheduler sử dụng.
        return Shift(
            key=self.key,
            label=self.label,
            start=parse_hhmm("00:00"),
            end=parse_hhmm("00:01"),
            minimum_minutes=0,
            excel_offset=0,
        )


@dataclass(frozen=True)
class Config:
    guild_id: int
    voice_channel_id: int
    report_channel_id: int
    timezone_name: str
    database_path: Path
    workbook_path: Path
    sheet_name: str
    header_row: int
    first_member_row: int
    member_end_marker: str
    stt_column: int
    attendance_columns: dict[str, tuple[int, ...]]
    scheduler_seconds: int
    manual_required_ratio: float
    command_role_ids: tuple[int, ...]
    shifts: tuple[Shift, ...]
    special_activities: tuple[SpecialActivity, ...]

    @property
    def tz(self) -> ZoneInfo:
        return ZoneInfo(self.timezone_name)

    def shift(self, key: str) -> Shift:
        for item in self.shifts:
            if item.key == key:
                return item
        for item in self.special_activities:
            if item.key == key:
                return item.as_shift()
        raise KeyError(f"Không tìm thấy ca/hoạt động {key!r}")

    def special_activity(self, key: str) -> SpecialActivity:
        for item in self.special_activities:
            if item.key == key:
                return item
        raise KeyError(f"Không tìm thấy hoạt động {key!r}")

    @property
    def fixed_shift_keys(self) -> set[str]:
        return {item.key for item in self.shifts}

    @property
    def all_activities(self) -> tuple[Shift, ...]:
        return self.shifts + tuple(item.as_shift() for item in self.special_activities)

    def attendance_column(self, shift_key: str, weekday: int) -> int:
        if weekday not in range(7):
            raise ValueError("weekday phải từ 0 (Thứ 2) đến 6 (Chủ nhật)")
        columns = self.attendance_columns.get(shift_key)
        if columns is None or len(columns) != 7:
            raise ValueError(f"Thiếu cấu hình 7 cột Excel cho ca {shift_key!r}")
        return int(columns[weekday])


def resolve_path(raw: str) -> Path:
    path = Path(raw)
    return path if path.is_absolute() else BASE_DIR / path


def load_config() -> Config:
    raw = json.loads(CONFIG_PATH.read_text(encoding="utf-8"))
    shifts = tuple(
        Shift(
            key=item["key"],
            label=item["label"],
            start=parse_hhmm(item["start"]),
            end=parse_hhmm(item["end"]),
            minimum_minutes=float(item["minimum_minutes"]),
            excel_offset=int(item["excel_offset"]),
        )
        for item in raw["shifts"]
    )
    special_activities = tuple(
        SpecialActivity(
            key=str(item["key"]),
            label=str(item["label"]),
            weekday=int(item["weekday"]),
            column=int(item["column"]),
        )
        for item in raw.get("special_activities", [])
    )
    for item in special_activities:
        if item.weekday not in range(7):
            raise ValueError(f"weekday của {item.key} phải từ 0 đến 6")
        if item.column <= 0:
            raise ValueError(f"column của {item.key} phải lớn hơn 0")
    ratio = float(raw.get("manual_required_ratio", 0.6))
    if not 0 < ratio <= 1:
        raise ValueError("manual_required_ratio phải lớn hơn 0 và không quá 1")
    attendance_columns = {
        str(key): tuple(int(column) for column in columns)
        for key, columns in raw.get("attendance_columns", {}).items()
    }
    all_keys = [shift.key for shift in shifts] + [item.key for item in special_activities]
    for key in all_keys:
        columns = attendance_columns.get(key)
        if columns is None or len(columns) != 7:
            raise ValueError(
                f"attendance_columns.{key} phải có đúng 7 giá trị cột"
            )
    for item in special_activities:
        if any(column != item.column for column in attendance_columns[item.key]):
            raise ValueError(
                f"attendance_columns.{item.key} phải lặp lại cột {item.column} đủ 7 lần"
            )

    return Config(
        guild_id=int(raw["guild_id"]),
        voice_channel_id=int(raw["voice_channel_id"]),
        report_channel_id=int(raw["report_channel_id"]),
        timezone_name=raw.get("timezone", "Asia/Ho_Chi_Minh"),
        database_path=resolve_path(raw.get("database_path", "data/attendance.sqlite3")),
        workbook_path=resolve_path(raw.get("workbook_path", "data/Tuyet_Nguyet.xlsx")),
        sheet_name=raw.get("sheet_name", "Trang tính2"),
        header_row=int(raw.get("header_row", 2)),
        first_member_row=int(raw.get("first_member_row", 3)),
        member_end_marker=str(
            raw.get("member_end_marker", "Số Người Tham Gia Hoạt Động")
        ).strip(),
        stt_column=int(raw.get("stt_column", 1)),
        attendance_columns=attendance_columns,
        scheduler_seconds=max(2, int(raw.get("scheduler_seconds", 5))),
        manual_required_ratio=ratio,
        command_role_ids=tuple(int(value) for value in raw.get("command_role_ids", [])),
        shifts=shifts,
        special_activities=special_activities,
    )


def ensure_workbook_seed(config: Config) -> None:
    config.workbook_path.parent.mkdir(parents=True, exist_ok=True)
    if not TEMPLATE_WORKBOOK_PATH.exists():
        raise FileNotFoundError(f"Không tìm thấy Excel mẫu: {TEMPLATE_WORKBOOK_PATH}")

    if not config.workbook_path.exists():
        shutil.copy2(TEMPLATE_WORKBOOK_PATH, config.workbook_path)
        log.info("Đã tạo Excel làm việc tại %s", config.workbook_path)
        return

    # Tự chuyển từ form cũ sang form mới. File cũ luôn được sao lưu trước.
    try:
        existing = load_workbook(config.workbook_path, read_only=True, data_only=False)
        compatible = config.sheet_name in existing.sheetnames
        existing.close()
    except Exception:
        compatible = False

    if compatible:
        return

    backup_dir = config.workbook_path.parent / "backups"
    backup_dir.mkdir(parents=True, exist_ok=True)
    backup = backup_dir / f"form_cu_{datetime.now():%Y%m%d_%H%M%S}.xlsx"
    shutil.copy2(config.workbook_path, backup)
    shutil.copy2(TEMPLATE_WORKBOOK_PATH, config.workbook_path)
    log.warning(
        "Excel cũ không có sheet %r; đã sao lưu tại %s và thay bằng form mới.",
        config.sheet_name,
        backup,
    )


def as_utc_text(value: datetime) -> str:
    if value.tzinfo is None:
        raise ValueError("datetime phải có múi giờ")
    return value.astimezone(UTC).isoformat()


def parse_dt(value: str | None) -> datetime | None:
    return datetime.fromisoformat(value) if value else None


class Database:
    def __init__(self, path: Path):
        self.path = path
        self.path.parent.mkdir(parents=True, exist_ok=True)
        self.lock = asyncio.Lock()

    def connect(self) -> sqlite3.Connection:
        connection = sqlite3.connect(self.path)
        connection.row_factory = sqlite3.Row
        connection.execute("PRAGMA journal_mode=WAL")
        connection.execute("PRAGMA foreign_keys=ON")
        return connection

    async def initialize(self) -> None:
        async with self.lock:
            with self.connect() as db:
                db.executescript(
                    """
                    CREATE TABLE IF NOT EXISTS shift_runs (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        shift_date TEXT NOT NULL,
                        shift_key TEXT NOT NULL,
                        shift_label TEXT NOT NULL,
                        mode TEXT NOT NULL CHECK(mode IN ('fixed', 'manual')),
                        started_at_utc TEXT NOT NULL,
                        scheduled_end_utc TEXT,
                        ended_at_utc TEXT,
                        required_seconds REAL,
                        status TEXT NOT NULL CHECK(status IN ('active', 'finalized', 'cancelled', 'missed')),
                        created_by_user_id INTEGER,
                        created_by_name TEXT
                    );

                    CREATE UNIQUE INDEX IF NOT EXISTS uq_active_shift_per_day
                    ON shift_runs(shift_date, shift_key)
                    WHERE status='active';

                    CREATE UNIQUE INDEX IF NOT EXISTS uq_fixed_shift_per_day
                    ON shift_runs(shift_date, shift_key, mode)
                    WHERE mode='fixed';

                    CREATE INDEX IF NOT EXISTS idx_shift_runs_status
                    ON shift_runs(status, shift_date, shift_key);

                    CREATE TABLE IF NOT EXISTS run_sessions (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        run_id INTEGER NOT NULL REFERENCES shift_runs(id) ON DELETE CASCADE,
                        user_id INTEGER NOT NULL,
                        display_name TEXT NOT NULL,
                        stt INTEGER,
                        started_at_utc TEXT NOT NULL,
                        ended_at_utc TEXT
                    );

                    CREATE UNIQUE INDEX IF NOT EXISTS uq_open_session_per_run_user
                    ON run_sessions(run_id, user_id)
                    WHERE ended_at_utc IS NULL;

                    CREATE INDEX IF NOT EXISTS idx_run_sessions_run
                    ON run_sessions(run_id, started_at_utc, ended_at_utc);

                    CREATE TABLE IF NOT EXISTS daily_overrides (
                        shift_date TEXT NOT NULL,
                        shift_key TEXT NOT NULL,
                        manual_run_id INTEGER NOT NULL REFERENCES shift_runs(id),
                        created_at_utc TEXT NOT NULL,
                        PRIMARY KEY(shift_date, shift_key)
                    );

                    CREATE TABLE IF NOT EXISTS bot_meta (
                        key TEXT PRIMARY KEY,
                        value TEXT NOT NULL
                    );
                    """
                )

    async def heartbeat(self, now: datetime) -> None:
        async with self.lock:
            with self.connect() as db:
                db.execute(
                    """
                    INSERT INTO bot_meta(key, value) VALUES ('last_heartbeat_utc', ?)
                    ON CONFLICT(key) DO UPDATE SET value=excluded.value
                    """,
                    (as_utc_text(now),),
                )

    async def recover_open_sessions(self, now: datetime) -> int:
        """Đóng phiên treo tại heartbeat cuối, không tính thời gian bot bị offline."""
        async with self.lock:
            with self.connect() as db:
                meta = db.execute(
                    "SELECT value FROM bot_meta WHERE key='last_heartbeat_utc'"
                ).fetchone()
                heartbeat = parse_dt(meta["value"]) if meta else None
                rows = db.execute(
                    "SELECT id, started_at_utc FROM run_sessions WHERE ended_at_utc IS NULL"
                ).fetchall()
                recovered = 0
                for row in rows:
                    started = datetime.fromisoformat(row["started_at_utc"])
                    close_at = min(now.astimezone(UTC), heartbeat or started)
                    if close_at > started:
                        db.execute(
                            "UPDATE run_sessions SET ended_at_utc=? WHERE id=?",
                            (close_at.isoformat(), int(row["id"])),
                        )
                    else:
                        db.execute("DELETE FROM run_sessions WHERE id=?", (int(row["id"]),))
                    recovered += 1
                return recovered

    async def active_runs(self) -> list[sqlite3.Row]:
        async with self.lock:
            with self.connect() as db:
                return db.execute(
                    "SELECT * FROM shift_runs WHERE status='active' ORDER BY id"
                ).fetchall()

    async def get_run(self, run_id: int) -> sqlite3.Row | None:
        async with self.lock:
            with self.connect() as db:
                return db.execute("SELECT * FROM shift_runs WHERE id=?", (run_id,)).fetchone()

    async def active_run_for_key(self, shift_key: str) -> sqlite3.Row | None:
        async with self.lock:
            with self.connect() as db:
                return db.execute(
                    """
                    SELECT * FROM shift_runs
                    WHERE shift_key=? AND status='active'
                    ORDER BY id DESC LIMIT 1
                    """,
                    (shift_key,),
                ).fetchone()

    async def fixed_run_exists(self, shift_day: date, shift_key: str) -> bool:
        async with self.lock:
            with self.connect() as db:
                row = db.execute(
                    """
                    SELECT 1 FROM shift_runs
                    WHERE shift_date=? AND shift_key=? AND mode='fixed'
                    """,
                    (shift_day.isoformat(), shift_key),
                ).fetchone()
                return row is not None

    async def has_override(self, shift_day: date, shift_key: str) -> bool:
        async with self.lock:
            with self.connect() as db:
                row = db.execute(
                    "SELECT 1 FROM daily_overrides WHERE shift_date=? AND shift_key=?",
                    (shift_day.isoformat(), shift_key),
                ).fetchone()
                return row is not None

    async def start_fixed_run(
        self,
        shift_day: date,
        shift: Shift,
        actual_start: datetime,
        scheduled_end: datetime,
    ) -> int | None:
        async with self.lock:
            with self.connect() as db:
                if db.execute(
                    "SELECT 1 FROM daily_overrides WHERE shift_date=? AND shift_key=?",
                    (shift_day.isoformat(), shift.key),
                ).fetchone():
                    return None
                if db.execute(
                    """
                    SELECT 1 FROM shift_runs
                    WHERE shift_date=? AND shift_key=? AND mode='fixed'
                    """,
                    (shift_day.isoformat(), shift.key),
                ).fetchone():
                    return None
                cursor = db.execute(
                    """
                    INSERT INTO shift_runs(
                        shift_date, shift_key, shift_label, mode,
                        started_at_utc, scheduled_end_utc,
                        required_seconds, status
                    ) VALUES (?, ?, ?, 'fixed', ?, ?, ?, 'active')
                    """,
                    (
                        shift_day.isoformat(),
                        shift.key,
                        shift.label,
                        as_utc_text(actual_start),
                        as_utc_text(scheduled_end),
                        shift.minimum_minutes * 60.0,
                    ),
                )
                return int(cursor.lastrowid)

    async def mark_fixed_missed(
        self, shift_day: date, shift: Shift, start: datetime, end: datetime
    ) -> None:
        async with self.lock:
            with self.connect() as db:
                if db.execute(
                    "SELECT 1 FROM daily_overrides WHERE shift_date=? AND shift_key=?",
                    (shift_day.isoformat(), shift.key),
                ).fetchone():
                    return
                if db.execute(
                    """
                    SELECT 1 FROM shift_runs
                    WHERE shift_date=? AND shift_key=? AND mode='fixed'
                    """,
                    (shift_day.isoformat(), shift.key),
                ).fetchone():
                    return
                db.execute(
                    """
                    INSERT INTO shift_runs(
                        shift_date, shift_key, shift_label, mode,
                        started_at_utc, scheduled_end_utc, ended_at_utc,
                        required_seconds, status
                    ) VALUES (?, ?, ?, 'fixed', ?, ?, ?, ?, 'missed')
                    """,
                    (
                        shift_day.isoformat(),
                        shift.key,
                        shift.label,
                        as_utc_text(start),
                        as_utc_text(end),
                        as_utc_text(end),
                        shift.minimum_minutes * 60.0,
                    ),
                )

    async def start_manual_run(
        self,
        shift_day: date,
        shift: Shift,
        started_at: datetime,
        user_id: int,
        user_name: str,
    ) -> tuple[int, int | None]:
        """Mở ca linh động và hủy ca cố định cùng loại nếu ca đó đang chạy."""
        async with self.lock:
            with self.connect() as db:
                existing = db.execute(
                    """
                    SELECT id FROM shift_runs
                    WHERE shift_key=? AND status='active' AND mode='manual'
                    ORDER BY id DESC LIMIT 1
                    """,
                    (shift.key,),
                ).fetchone()
                if existing:
                    raise ValueError(f"Ca {shift.label} linh động đang chạy rồi.")

                cancelled_id: int | None = None
                fixed = db.execute(
                    """
                    SELECT id FROM shift_runs
                    WHERE shift_date=? AND shift_key=? AND status='active' AND mode='fixed'
                    ORDER BY id DESC LIMIT 1
                    """,
                    (shift_day.isoformat(), shift.key),
                ).fetchone()
                if fixed:
                    cancelled_id = int(fixed["id"])
                    db.execute(
                        """
                        UPDATE run_sessions SET ended_at_utc=?
                        WHERE run_id=? AND ended_at_utc IS NULL
                        """,
                        (as_utc_text(started_at), cancelled_id),
                    )
                    db.execute(
                        """
                        UPDATE shift_runs
                        SET status='cancelled', ended_at_utc=?
                        WHERE id=?
                        """,
                        (as_utc_text(started_at), cancelled_id),
                    )

                cursor = db.execute(
                    """
                    INSERT INTO shift_runs(
                        shift_date, shift_key, shift_label, mode,
                        started_at_utc, status,
                        created_by_user_id, created_by_name
                    ) VALUES (?, ?, ?, 'manual', ?, 'active', ?, ?)
                    """,
                    (
                        shift_day.isoformat(),
                        shift.key,
                        shift.label,
                        as_utc_text(started_at),
                        user_id,
                        user_name,
                    ),
                )
                run_id = int(cursor.lastrowid)
                db.execute(
                    """
                    INSERT INTO daily_overrides(shift_date, shift_key, manual_run_id, created_at_utc)
                    VALUES (?, ?, ?, ?)
                    ON CONFLICT(shift_date, shift_key) DO UPDATE SET
                        manual_run_id=excluded.manual_run_id,
                        created_at_utc=excluded.created_at_utc
                    """,
                    (shift_day.isoformat(), shift.key, run_id, as_utc_text(started_at)),
                )
                return run_id, cancelled_id

    async def begin_member(
        self,
        run_ids: list[int],
        user_id: int,
        display_name: str,
        stt: int | None,
        started_at: datetime,
    ) -> None:
        if not run_ids:
            return
        async with self.lock:
            with self.connect() as db:
                for run_id in run_ids:
                    db.execute(
                        """
                        INSERT OR IGNORE INTO run_sessions(
                            run_id, user_id, display_name, stt, started_at_utc
                        ) VALUES (?, ?, ?, ?, ?)
                        """,
                        (run_id, user_id, display_name, stt, as_utc_text(started_at)),
                    )

    async def end_member(self, run_ids: list[int], user_id: int, ended_at: datetime) -> None:
        if not run_ids:
            return
        async with self.lock:
            with self.connect() as db:
                placeholders = ",".join("?" for _ in run_ids)
                db.execute(
                    f"""
                    UPDATE run_sessions SET ended_at_utc=?
                    WHERE user_id=? AND ended_at_utc IS NULL
                      AND run_id IN ({placeholders})
                    """,
                    (as_utc_text(ended_at), user_id, *run_ids),
                )

    async def attendance_at(
        self, run_id: int, ended_at: datetime
    ) -> dict[int, dict[str, Any]]:
        end_utc = ended_at.astimezone(UTC)
        async with self.lock:
            with self.connect() as db:
                rows = db.execute(
                    """
                    SELECT user_id, display_name, stt, started_at_utc, ended_at_utc
                    FROM run_sessions WHERE run_id=?
                    """,
                    (run_id,),
                ).fetchall()

        totals: dict[int, float] = {}
        names: dict[int, str] = {}
        stts: dict[int, int | None] = {}
        for row in rows:
            started = datetime.fromisoformat(row["started_at_utc"])
            stored_end = parse_dt(row["ended_at_utc"])
            segment_end = min(stored_end or end_utc, end_utc)
            seconds = max(0.0, (segment_end - started).total_seconds())
            if seconds <= 0:
                continue
            user_id = int(row["user_id"])
            totals[user_id] = totals.get(user_id, 0.0) + seconds
            names[user_id] = row["display_name"]
            stts[user_id] = int(row["stt"]) if row["stt"] is not None else None
        return {
            user_id: {
                "display_name": names[user_id],
                "stt": stts[user_id],
                "seconds": seconds,
            }
            for user_id, seconds in totals.items()
        }

    async def finalize_run(self, run_id: int, ended_at: datetime, required_seconds: float) -> None:
        async with self.lock:
            with self.connect() as db:
                db.execute(
                    """
                    UPDATE run_sessions SET ended_at_utc=?
                    WHERE run_id=? AND ended_at_utc IS NULL
                    """,
                    (as_utc_text(ended_at), run_id),
                )
                db.execute(
                    """
                    UPDATE shift_runs
                    SET ended_at_utc=?, required_seconds=?, status='finalized'
                    WHERE id=? AND status='active'
                    """,
                    (as_utc_text(ended_at), required_seconds, run_id),
                )

    async def cancel_run(self, run_id: int, ended_at: datetime, remove_override: bool = False) -> None:
        """Hủy ca khi bot không thể vào voice; không để lại ca active bị kẹt."""
        async with self.lock:
            with self.connect() as db:
                run = db.execute(
                    "SELECT shift_date, shift_key FROM shift_runs WHERE id=?",
                    (run_id,),
                ).fetchone()
                db.execute(
                    """
                    UPDATE run_sessions SET ended_at_utc=?
                    WHERE run_id=? AND ended_at_utc IS NULL
                    """,
                    (as_utc_text(ended_at), run_id),
                )
                db.execute(
                    """
                    UPDATE shift_runs SET status='cancelled', ended_at_utc=?
                    WHERE id=? AND status='active'
                    """,
                    (as_utc_text(ended_at), run_id),
                )
                if remove_override and run is not None:
                    db.execute(
                        """
                        DELETE FROM daily_overrides
                        WHERE shift_date=? AND shift_key=? AND manual_run_id=?
                        """,
                        (run["shift_date"], run["shift_key"], run_id),
                    )

    async def clear_week(self, monday: date) -> None:
        sunday = monday + timedelta(days=6)
        async with self.lock:
            with self.connect() as db:
                db.execute(
                    "DELETE FROM daily_overrides WHERE shift_date BETWEEN ? AND ?",
                    (monday.isoformat(), sunday.isoformat()),
                )
                db.execute(
                    "DELETE FROM shift_runs WHERE shift_date BETWEEN ? AND ?",
                    (monday.isoformat(), sunday.isoformat()),
                )


class WorkbookManager:
    def __init__(self, config: Config):
        self.config = config
        self.lock = asyncio.Lock()
        self.backup_dir = self.config.workbook_path.parent / "backups"
        self.export_dir = self.config.workbook_path.parent / "exports"
        self.backup_dir.mkdir(parents=True, exist_ok=True)
        self.export_dir.mkdir(parents=True, exist_ok=True)

    @staticmethod
    def _normalize_text(value: Any) -> str:
        text = str(value or "").strip().casefold()
        normalized = "".join(
            char
            for char in unicodedata.normalize("NFD", text)
            if unicodedata.category(char) != "Mn"
        )
        return normalized.replace("đ", "d")

    def _row_by_stt(self, worksheet: Any) -> dict[int, int]:
        """Đọc động toàn bộ thành viên cho tới dòng tổng kết.

        Người dùng có thể chèn thêm bao nhiêu dòng thành viên tùy ý phía trên
        dòng 'Số Người Tham Gia Hoạt Động'; bot không phụ thuộc số hàng cố định.
        """
        result: dict[int, int] = {}
        marker = self._normalize_text(self.config.member_end_marker)
        for row in range(self.config.first_member_row, worksheet.max_row + 1):
            value = worksheet.cell(row=row, column=self.config.stt_column).value
            if marker and self._normalize_text(value) == marker:
                break
            if isinstance(value, bool):
                continue
            try:
                stt = int(float(value))
            except (TypeError, ValueError):
                continue
            if stt > 0:
                if stt in result:
                    raise ValueError(
                        f"STT {stt} bị trùng ở hàng {result[stt]} và hàng {row}."
                    )
                result[stt] = row
        return result

    def validate_path_sync(self, workbook_path: Path) -> dict[str, Any]:
        if not workbook_path.exists():
            raise FileNotFoundError(f"Không tìm thấy file Excel: {workbook_path}")
        workbook = load_workbook(workbook_path, data_only=False)
        if self.config.sheet_name not in workbook.sheetnames:
            raise ValueError(f"Không tìm thấy sheet {self.config.sheet_name!r}")
        worksheet = workbook[self.config.sheet_name]
        rows = self._row_by_stt(worksheet)
        if not rows:
            raise ValueError("Không tìm thấy STT trong file Excel.")
        expected_tokens = {
            "boss_trua": ("12h",),
            "van_tieu": ("tieu",),
            "boss_toi": ("10h", "22h"),
            "lien_dai_thu2": ("lien dai",),
            "lien_dai_thu5": ("lien dai",),
            "ctc": ("ctc",),
        }
        for shift in self.config.all_activities:
            for weekday in range(7):
                column = self.config.attendance_column(shift.key, weekday)
                actual = worksheet.cell(
                    row=self.config.header_row, column=column
                ).value
                normalized = self._normalize_text(actual)
                tokens = expected_tokens.get(shift.key, ())
                if not normalized or (tokens and not any(token in normalized for token in tokens)):
                    raise ValueError(
                        f"Excel sai cấu trúc tại "
                        f"{get_column_letter(column)}{self.config.header_row}: "
                        f"không nhận ra cột {shift.label!r}, đang là {actual!r}."
                    )
        return {"sheet": worksheet.title, "members": len(rows), "max_stt": max(rows)}

    def validate_sync(self) -> dict[str, Any]:
        return self.validate_path_sync(self.config.workbook_path)

    async def validate(self) -> dict[str, Any]:
        return await asyncio.to_thread(self.validate_sync)

    def install_uploaded_workbook_sync(self, uploaded_path: Path) -> tuple[Path, dict[str, Any]]:
        """Kiểm tra form, sao lưu file hiện tại rồi thay bằng Excel được tải lên."""
        info = self.validate_path_sync(uploaded_path)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup = self.backup_dir / f"before_manual_excel_update_{timestamp}.xlsx"
        if self.config.workbook_path.exists():
            shutil.copy2(self.config.workbook_path, backup)
        else:
            backup = Path("")

        replacement = self.config.workbook_path.with_name(
            f".{self.config.workbook_path.stem}_upload_{timestamp}.tmp.xlsx"
        )
        shutil.copy2(uploaded_path, replacement)
        os.replace(replacement, self.config.workbook_path)
        return backup, info

    async def install_uploaded_workbook(self, uploaded_path: Path) -> tuple[Path, dict[str, Any]]:
        async with self.lock:
            return await asyncio.to_thread(self.install_uploaded_workbook_sync, uploaded_path)

    def replace_attendance_sync(
        self,
        shift_day: date,
        shift: Shift,
        qualified: list[tuple[int, str, float]],
    ) -> tuple[Path, list[tuple[int, str, float]], list[tuple[str, float]]]:
        """Xóa kết quả cũ của đúng cột rồi ghi kết quả ca mới vào."""
        self.validate_sync()
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup = self.backup_dir / (
            f"before_{shift_day.isoformat()}_{shift.key}_{timestamp}.xlsx"
        )
        shutil.copy2(self.config.workbook_path, backup)

        workbook = load_workbook(self.config.workbook_path, data_only=False)
        worksheet = workbook[self.config.sheet_name]
        rows = self._row_by_stt(worksheet)
        target_column = self.config.attendance_column(
            shift.key, shift_day.weekday()
        )

        # Quan trọng: ca linh động thay thế hoàn toàn ca cố định của ngày đó.
        for row in rows.values():
            worksheet.cell(row=row, column=target_column).value = None

        recorded: list[tuple[int, str, float]] = []
        missing: list[tuple[str, float]] = []
        for stt, name, seconds in qualified:
            row = rows.get(stt)
            if row is None:
                missing.append((name, seconds))
                continue
            worksheet.cell(row=row, column=target_column).value = 1
            recorded.append((stt, name, seconds))

        try:
            workbook.calculation.calcMode = "auto"
            workbook.calculation.fullCalcOnLoad = True
            workbook.calculation.forceFullCalc = True
        except AttributeError:
            pass

        temp_path = self.config.workbook_path.with_name(
            f".{self.config.workbook_path.stem}_{timestamp}.tmp.xlsx"
        )
        workbook.save(temp_path)
        try:
            os.replace(temp_path, self.config.workbook_path)
            saved_path = self.config.workbook_path
        except PermissionError:
            saved_path = self.export_dir / (
                f"Tuyet_Nguyet_{shift_day.isoformat()}_{shift.key}_{timestamp}.xlsx"
            )
            shutil.move(temp_path, saved_path)
        return saved_path, recorded, missing

    async def replace_attendance(
        self,
        shift_day: date,
        shift: Shift,
        qualified: list[tuple[int, str, float]],
    ) -> tuple[Path, list[tuple[int, str, float]], list[tuple[str, float]]]:
        async with self.lock:
            return await asyncio.to_thread(
                self.replace_attendance_sync, shift_day, shift, qualified
            )

    def new_week_sync(self, monday: date) -> Path:
        self.validate_sync()
        archive_dir = self.config.workbook_path.parent / "archives"
        archive_dir.mkdir(parents=True, exist_ok=True)
        archive_path = archive_dir / f"Tuyet_Nguyet_tuan_{monday.isoformat()}.xlsx"
        if archive_path.exists():
            archive_path = archive_dir / (
                f"Tuyet_Nguyet_tuan_{monday.isoformat()}_{datetime.now():%H%M%S}.xlsx"
            )
        shutil.copy2(self.config.workbook_path, archive_path)

        workbook = load_workbook(self.config.workbook_path, data_only=False)
        worksheet = workbook[self.config.sheet_name]
        rows = self._row_by_stt(worksheet)
        attendance_columns = sorted(
            {
                column
                for columns in self.config.attendance_columns.values()
                for column in columns
            }
        )
        for row in rows.values():
            for column in attendance_columns:
                worksheet.cell(row=row, column=column).value = None
        workbook.save(self.config.workbook_path)
        return archive_path

    async def new_week(self, monday: date) -> Path:
        async with self.lock:
            return await asyncio.to_thread(self.new_week_sync, monday)


CONFIG = load_config()
ensure_workbook_seed(CONFIG)
DB = Database(CONFIG.database_path)
WORKBOOK = WorkbookManager(CONFIG)

intents = discord.Intents.none()
intents.guilds = True
intents.voice_states = True
client = discord.Client(intents=intents)
tree = app_commands.CommandTree(client)
RUN_LOCK = asyncio.Lock()
_bootstrapped = False
_commands_synced = False


def weekday_text(value: date) -> str:
    names = ["Thứ 2", "Thứ 3", "Thứ 4", "Thứ 5", "Thứ 6", "Thứ 7", "Chủ nhật"]
    return names[value.weekday()]


def can_control(interaction: discord.Interaction) -> bool:
    member = interaction.user
    if not isinstance(member, discord.Member):
        return False
    permissions = member.guild_permissions
    if permissions.administrator or permissions.manage_guild or permissions.manage_channels:
        return True
    allowed_roles = set(CONFIG.command_role_ids)
    return bool(allowed_roles and any(role.id in allowed_roles for role in member.roles))


def command_is_in_voice_chat(interaction: discord.Interaction) -> bool:
    member = interaction.user
    return (
        interaction.guild_id == CONFIG.guild_id
        and interaction.channel_id == CONFIG.voice_channel_id
        and isinstance(member, discord.Member)
        and member.voice is not None
        and member.voice.channel is not None
        and member.voice.channel.id == CONFIG.voice_channel_id
    )


async def resolve_voice_channel() -> discord.VoiceChannel:
    guild = client.get_guild(CONFIG.guild_id)
    if guild is None:
        raise RuntimeError("Bot không tìm thấy server đã cấu hình.")
    channel = guild.get_channel(CONFIG.voice_channel_id)
    if not isinstance(channel, discord.VoiceChannel):
        raise RuntimeError("Voice Channel ID sai hoặc bot không có quyền View Channel.")
    return channel


async def resolve_report_channel() -> discord.abc.Messageable:
    guild = client.get_guild(CONFIG.guild_id)
    if guild is None:
        raise RuntimeError("Bot không tìm thấy server đã cấu hình.")
    channel = guild.get_channel(CONFIG.report_channel_id)
    if not isinstance(channel, (discord.TextChannel, discord.VoiceChannel)):
        raise RuntimeError("Report Channel ID sai hoặc bot không xem được kênh.")
    return channel


async def voice_members(channel: discord.VoiceChannel) -> list[discord.Member]:
    members: list[discord.Member] = []
    for user_id in channel.voice_states:
        member = channel.guild.get_member(user_id)
        if member is None:
            try:
                member = await channel.guild.fetch_member(user_id)
            except discord.HTTPException:
                continue
        if not member.bot:
            members.append(member)
    return members


async def ensure_bot_in_voice() -> None:
    channel = await resolve_voice_channel()
    voice_client = channel.guild.voice_client
    if voice_client and voice_client.is_connected():
        if voice_client.channel.id != channel.id:
            await voice_client.move_to(channel)
        return
    if voice_client:
        try:
            await voice_client.disconnect(force=True)
        except Exception:
            log.exception("Không dọn được voice connection cũ")
    await channel.connect(self_deaf=True, self_mute=True, reconnect=True, timeout=30)
    log.info("Bot đã vào voice: %s", channel.name)


async def disconnect_bot_if_idle() -> None:
    if await DB.active_runs():
        return
    guild = client.get_guild(CONFIG.guild_id)
    voice_client = guild.voice_client if guild else None
    if voice_client and voice_client.is_connected():
        await voice_client.disconnect(force=False)
        log.info("Bot đã rời voice vì không còn ca đang chạy.")


async def seed_members_for_run(run_id: int, started_at: datetime) -> int:
    channel = await resolve_voice_channel()
    members = await voice_members(channel)
    for member in members:
        await DB.begin_member(
            [run_id],
            member.id,
            member.display_name,
            extract_stt(member.display_name),
            started_at,
        )
    return len(members)


async def announce_in_voice(message: str) -> None:
    try:
        channel = await resolve_voice_channel()
        await channel.send(message)
    except discord.HTTPException:
        log.exception("Không gửi được tin nhắn vào chat của kênh voice")


async def start_fixed_shift(shift_day: date, shift: Shift, now: datetime) -> None:
    start, end = shift_window(shift_day, shift, CONFIG.tz)
    actual_start = max(now, start)
    async with RUN_LOCK:
        run_id = await DB.start_fixed_run(shift_day, shift, actual_start, end)
        if run_id is None:
            return
        try:
            await ensure_bot_in_voice()
            count = await seed_members_for_run(run_id, actual_start.astimezone(UTC))
        except Exception:
            await DB.cancel_run(run_id, datetime.now(UTC))
            log.exception("Không thể bắt đầu ca cố định %s", shift.label)
            return
    log.info("Bắt đầu ca cố định %s, run_id=%s, %s người", shift.label, run_id, count)
    await announce_in_voice(
        f"🟢 **Bắt đầu {shift.label}** — {shift.start:%H:%M} đến {shift.end:%H:%M}.\n"
        f"Yêu cầu có mặt tối thiểu **{shift.minimum_minutes:g} phút**. Bot đang ngồi trong voice để điểm danh."
    )


async def build_and_send_report(
    run: sqlite3.Row,
    shift: Shift,
    ended_at: datetime,
    required_seconds: float,
    saved_path: Path,
    recorded: list[tuple[int, str, float]],
    insufficient: list[tuple[str, float]],
    no_stt: list[tuple[str, float]],
    missing_excel: list[tuple[str, float]],
) -> None:
    shift_day = date.fromisoformat(run["shift_date"])
    mode = "ca linh động" if run["mode"] == "manual" else "ca cố định"
    start = datetime.fromisoformat(run["started_at_utc"]).astimezone(CONFIG.tz)
    lines = [
        f"**Kết quả {shift.label} — {weekday_text(shift_day)} ({shift_day:%d/%m/%Y})**",
        f"Loại: **{mode}** | Thời gian: **{start:%H:%M:%S}–{ended_at.astimezone(CONFIG.tz):%H:%M:%S}**",
        f"Mức đạt: **{format_duration(required_seconds)}**",
        f"✅ Đã ghi `1` vào Excel: **{len(recorded)} người**",
    ]
    if run["mode"] == "manual" and run["shift_key"] in CONFIG.fixed_shift_keys:
        lines.append("🔄 Kết quả này **thay thế hoàn toàn** kết quả ca cố định cùng loại của ngày này.")
    elif run["mode"] == "manual":
        lines.append("📝 Đây là hoạt động mở bằng lệnh; kết quả mới thay thế kết quả cũ của đúng cột hoạt động này.")
    if recorded:
        lines.append(
            "• " + ", ".join(
                f"STT {stt} ({format_duration(seconds)})"
                for stt, _name, seconds in recorded
            )
        )
    if insufficient:
        lines.append(
            "⚠️ Chưa đủ thời gian: "
            + ", ".join(f"{name} ({format_duration(seconds)})" for name, seconds in insufficient)
        )
    if no_stt:
        lines.append("❌ Tên không có STT đầu tên: " + ", ".join(name for name, _ in no_stt))
    if missing_excel:
        lines.append("❌ Có STT nhưng không thấy trong Excel: " + ", ".join(name for name, _ in missing_excel))
    if saved_path != CONFIG.workbook_path:
        lines.append("⚠️ Excel gốc đang bị khóa; bot đã lưu thành file xuất riêng.")

    message = "\n".join(lines)
    if len(message) > 1900:
        message = message[:1840] + "\n…Danh sách quá dài, xem file Excel đính kèm."
    channel = await resolve_report_channel()
    await channel.send(message, file=discord.File(saved_path, filename="Tuyet_Nguyet.xlsx"))


async def finalize_run(run_id: int, requested_end: datetime | None = None) -> dict[str, Any]:
    async with RUN_LOCK:
        run = await DB.get_run(run_id)
        if run is None or run["status"] != "active":
            raise ValueError("Ca này không còn ở trạng thái đang chạy.")
        shift = CONFIG.shift(run["shift_key"])
        now_utc = datetime.now(UTC)
        ended_at = (requested_end or now_utc).astimezone(UTC)
        started_at = datetime.fromisoformat(run["started_at_utc"])
        if ended_at <= started_at:
            raise ValueError("Thời gian kết thúc phải sau thời gian bắt đầu.")

        if run["mode"] == "manual":
            duration_seconds = (ended_at - started_at).total_seconds()
            required_seconds = manual_required_seconds(
                duration_seconds, CONFIG.manual_required_ratio
            )
        else:
            required_seconds = float(run["required_seconds"])

        attendance = await DB.attendance_at(run_id, ended_at)
        qualified: list[tuple[int, str, float]] = []
        insufficient: list[tuple[str, float]] = []
        no_stt: list[tuple[str, float]] = []
        for item in sorted(attendance.values(), key=lambda value: value["display_name"].casefold()):
            name = str(item["display_name"])
            stt = item["stt"]
            seconds = float(item["seconds"])
            if seconds + 1e-6 < required_seconds:
                insufficient.append((name, seconds))
            elif stt is None:
                no_stt.append((name, seconds))
            else:
                qualified.append((int(stt), name, seconds))

        shift_day = date.fromisoformat(run["shift_date"])
        saved_path, recorded, missing_excel = await WORKBOOK.replace_attendance(
            shift_day, shift, qualified
        )
        await DB.finalize_run(run_id, ended_at, required_seconds)

    await build_and_send_report(
        run,
        shift,
        ended_at,
        required_seconds,
        saved_path,
        recorded,
        insufficient,
        no_stt,
        missing_excel,
    )
    await disconnect_bot_if_idle()
    log.info("Đã chốt %s run_id=%s", shift.label, run_id)
    return {
        "shift": shift,
        "required_seconds": required_seconds,
        "recorded": len(recorded),
        "ended_at": ended_at,
    }


async def start_manual_command(interaction: discord.Interaction, shift_key: str) -> None:
    if not can_control(interaction):
        await interaction.response.send_message(
            "Bạn cần quyền **Manage Server**, **Manage Channels**, Administrator hoặc role được cấu hình.",
            ephemeral=True,
        )
        return
    if not command_is_in_voice_chat(interaction):
        await interaction.response.send_message(
            "Hãy vào đúng room voice điểm danh rồi dùng lệnh trong phần chat của chính room đó.",
            ephemeral=True,
        )
        return

    shift = CONFIG.shift(shift_key)
    now_local = datetime.now(CONFIG.tz)
    await interaction.response.defer(thinking=True)
    run_id: int | None = None
    try:
        async with RUN_LOCK:
            run_id, cancelled_id = await DB.start_manual_run(
                now_local.date(),
                shift,
                now_local.astimezone(UTC),
                interaction.user.id,
                interaction.user.display_name,
            )
            try:
                await ensure_bot_in_voice()
                count = await seed_members_for_run(run_id, now_local.astimezone(UTC))
            except Exception:
                await DB.cancel_run(run_id, datetime.now(UTC), remove_override=True)
                raise
        cancelled_text = (
            " Ca cố định đang chạy đã được hủy để chuyển sang ca linh động."
            if cancelled_id
            else ""
        )
        await interaction.followup.send(
            f"🟢 **Đã bắt đầu {shift.label} linh động lúc {now_local:%H:%M:%S}.**\n"
            f"Bot đã vào voice và bắt đầu tính cho **{count} người** đang có mặt.{cancelled_text}\n"
            f"Khi kết thúc, dùng lệnh kết thúc tương ứng. Mức đạt bằng **3/5 tổng thời gian ca**. "
            f"Ca này sẽ thay thế cột {shift.label} của {weekday_text(now_local.date())}.",
        )
    except Exception as exc:
        log.exception("Không mở được ca linh động %s", shift.label)
        await interaction.followup.send(f"❌ Không mở được ca: {exc}", ephemeral=True)


async def end_manual_command(interaction: discord.Interaction, shift_key: str) -> None:
    if not can_control(interaction):
        await interaction.response.send_message(
            "Bạn không có quyền kết thúc ca.", ephemeral=True
        )
        return
    if not command_is_in_voice_chat(interaction):
        await interaction.response.send_message(
            "Hãy dùng lệnh trong phần chat của đúng room voice điểm danh.", ephemeral=True
        )
        return
    shift = CONFIG.shift(shift_key)
    run = await DB.active_run_for_key(shift_key)
    if run is None or run["mode"] != "manual":
        await interaction.response.send_message(
            f"Hiện không có ca {shift.label} linh động nào đang chạy.", ephemeral=True
        )
        return

    await interaction.response.defer(thinking=True)
    try:
        result = await finalize_run(int(run["id"]), datetime.now(UTC))
        await interaction.followup.send(
            f"🔴 **Đã kết thúc {shift.label} linh động.**\n"
            f"Mức đạt: **{format_duration(result['required_seconds'])}**. "
            f"Đã ghi `1` cho **{result['recorded']} người** và gửi file Excel vào kênh báo cáo."
        )
    except Exception as exc:
        log.exception("Không kết thúc được ca linh động %s", shift.label)
        await interaction.followup.send(f"❌ Không kết thúc được ca: {exc}", ephemeral=True)


async def toggle_special_activity_command(
    interaction: discord.Interaction, activity_key: str
) -> None:
    """Một lệnh dùng để mở hoặc kết thúc hoạt động đặc biệt.

    Lần gọi đầu: bắt đầu điểm danh, bot vào voice.
    Lần gọi tiếp theo: kết thúc, yêu cầu 3/5 thời lượng và ghi Excel.
    """
    if not can_control(interaction):
        await interaction.response.send_message(
            "Bạn cần quyền **Manage Server**, **Manage Channels**, Administrator hoặc role được cấu hình.",
            ephemeral=True,
        )
        return
    if not command_is_in_voice_chat(interaction):
        await interaction.response.send_message(
            "Hãy vào đúng room voice điểm danh rồi dùng lệnh trong phần chat của chính room đó.",
            ephemeral=True,
        )
        return

    activity = CONFIG.special_activity(activity_key)
    now_local = datetime.now(CONFIG.tz)
    active = await DB.active_run_for_key(activity_key)
    if active is not None:
        if active["mode"] != "manual":
            await interaction.response.send_message(
                f"{activity.label} đang có một ca khác chạy, không thể chốt bằng lệnh này.",
                ephemeral=True,
            )
            return
        await interaction.response.defer(thinking=True)
        try:
            result = await finalize_run(int(active["id"]), datetime.now(UTC))
            await interaction.followup.send(
                f"🔴 **Đã kết thúc {activity.label}.**\n"
                f"Mức đạt: **{format_duration(result['required_seconds'])}** (3/5 thời lượng). "
                f"Đã ghi `1` cho **{result['recorded']} người**."
            )
        except Exception as exc:
            log.exception("Không kết thúc được %s", activity.label)
            await interaction.followup.send(
                f"❌ Không kết thúc được hoạt động: {exc}", ephemeral=True
            )
        return

    if now_local.weekday() != activity.weekday:
        expected = ["Thứ 2", "Thứ 3", "Thứ 4", "Thứ 5", "Thứ 6", "Thứ 7", "Chủ nhật"][activity.weekday]
        await interaction.response.send_message(
            f"Lệnh này chỉ dùng để **bắt đầu** vào {expected}. Hôm nay là **{weekday_text(now_local.date())}**.",
            ephemeral=True,
        )
        return

    await interaction.response.defer(thinking=True)
    run_id: int | None = None
    try:
        async with RUN_LOCK:
            shift = activity.as_shift()
            run_id, _ = await DB.start_manual_run(
                now_local.date(),
                shift,
                now_local.astimezone(UTC),
                interaction.user.id,
                interaction.user.display_name,
            )
            try:
                await ensure_bot_in_voice()
                count = await seed_members_for_run(run_id, now_local.astimezone(UTC))
            except Exception:
                await DB.cancel_run(run_id, datetime.now(UTC), remove_override=True)
                raise
        await interaction.followup.send(
            f"🟢 **Đã bắt đầu {activity.label} lúc {now_local:%H:%M:%S}.**\n"
            f"Bot đã vào voice và bắt đầu tính cho **{count} người** đang có mặt.\n"
            f"Dùng lại chính lệnh này để kết thúc. Mức đạt bằng **3/5 tổng thời gian hoạt động**."
        )
    except Exception as exc:
        log.exception("Không mở được %s", activity.label)
        await interaction.followup.send(
            f"❌ Không mở được hoạt động: {exc}", ephemeral=True
        )


@client.event
async def on_ready() -> None:
    global _bootstrapped, _commands_synced
    assert client.user is not None
    log.info("Đã đăng nhập: %s (%s)", client.user, client.user.id)

    guild = client.get_guild(CONFIG.guild_id)
    if guild is None:
        log.error("Bot chưa có trong Server ID %s", CONFIG.guild_id)
        return
    await resolve_voice_channel()
    await resolve_report_channel()

    if not _commands_synced:
        await tree.sync(guild=discord.Object(id=CONFIG.guild_id))
        _commands_synced = True
        log.info("Đã đồng bộ slash commands.")

    if not _bootstrapped:
        recovered = await DB.recover_open_sessions(datetime.now(UTC))
        active = await DB.active_runs()
        if active:
            await ensure_bot_in_voice()
            now = datetime.now(UTC)
            channel = await resolve_voice_channel()
            members = await voice_members(channel)
            run_ids = [int(run["id"]) for run in active]
            for member in members:
                await DB.begin_member(
                    run_ids,
                    member.id,
                    member.display_name,
                    extract_stt(member.display_name),
                    now,
                )
        _bootstrapped = True
        log.info("Khôi phục %s phiên treo; có %s ca đang chạy.", recovered, len(active))
        if not scheduler.is_running():
            scheduler.change_interval(seconds=CONFIG.scheduler_seconds)
            scheduler.start()


@client.event
async def on_voice_state_update(
    member: discord.Member,
    before: discord.VoiceState,
    after: discord.VoiceState,
) -> None:
    if member.bot or member.guild.id != CONFIG.guild_id:
        return
    before_target = before.channel is not None and before.channel.id == CONFIG.voice_channel_id
    after_target = after.channel is not None and after.channel.id == CONFIG.voice_channel_id
    if before_target == after_target:
        return
    async with RUN_LOCK:
        active = await DB.active_runs()
        run_ids = [int(run["id"]) for run in active]
        if not run_ids:
            return
        now = datetime.now(UTC)
        if after_target:
            await DB.begin_member(
                run_ids,
                member.id,
                member.display_name,
                extract_stt(member.display_name),
                now,
            )
            log.info("JOIN | %s | ca=%s", member.display_name, run_ids)
        else:
            await DB.end_member(run_ids, member.id, now)
            log.info("LEAVE | %s | ca=%s", member.display_name, run_ids)


@tasks.loop(seconds=5)
async def scheduler() -> None:
    now_utc = datetime.now(UTC)
    await DB.heartbeat(now_utc)
    now = now_utc.astimezone(CONFIG.tz)

    # Chốt ca cố định đã tới giờ kết thúc.
    for run in await DB.active_runs():
        if run["mode"] != "fixed" or not run["scheduled_end_utc"]:
            continue
        scheduled_end = datetime.fromisoformat(run["scheduled_end_utc"])
        if now_utc >= scheduled_end:
            try:
                await finalize_run(int(run["id"]), scheduled_end)
            except Exception:
                log.exception("Lỗi chốt ca cố định run_id=%s", run["id"])

    # Khởi động ca cố định hiện tại; ca đã có lệnh linh động trong ngày sẽ bị bỏ qua.
    for shift_day in (now.date() - timedelta(days=1), now.date()):
        for shift in CONFIG.shifts:
            start, end = shift_window(shift_day, shift, CONFIG.tz)
            if start <= now < end:
                if await DB.has_override(shift_day, shift.key):
                    continue
                if not await DB.fixed_run_exists(shift_day, shift.key):
                    try:
                        await start_fixed_shift(shift_day, shift, now)
                    except Exception:
                        log.exception("Lỗi mở ca cố định %s", shift.label)
            elif now >= end and not await DB.fixed_run_exists(shift_day, shift.key):
                # Bot đã offline toàn bộ ca: ghi trạng thái missed nhưng không gửi báo cáo rỗng.
                await DB.mark_fixed_missed(shift_day, shift, start, end)


@scheduler.before_loop
async def before_scheduler() -> None:
    await client.wait_until_ready()


GUILD_SCOPE = discord.Object(id=CONFIG.guild_id)


@tree.command(name="vantieu", description="Bắt đầu ca Vận Tiêu linh động", guild=GUILD_SCOPE)
async def vantieu(interaction: discord.Interaction) -> None:
    await start_manual_command(interaction, "van_tieu")


@tree.command(name="ketthucvantieu", description="Kết thúc ca Vận Tiêu linh động", guild=GUILD_SCOPE)
async def ketthucvantieu(interaction: discord.Interaction) -> None:
    await end_manual_command(interaction, "van_tieu")


@tree.command(name="bosstrua", description="Bắt đầu ca Boss Trưa linh động", guild=GUILD_SCOPE)
async def bosstrua(interaction: discord.Interaction) -> None:
    await start_manual_command(interaction, "boss_trua")


@tree.command(name="ketthucbosstrua", description="Kết thúc ca Boss Trưa linh động", guild=GUILD_SCOPE)
async def ketthucbosstrua(interaction: discord.Interaction) -> None:
    await end_manual_command(interaction, "boss_trua")


@tree.command(name="bosstoi", description="Bắt đầu ca Boss Tối linh động", guild=GUILD_SCOPE)
async def bosstoi(interaction: discord.Interaction) -> None:
    await start_manual_command(interaction, "boss_toi")


@tree.command(name="ketthucbosstoi", description="Kết thúc ca Boss Tối linh động", guild=GUILD_SCOPE)
async def ketthucbosstoi(interaction: discord.Interaction) -> None:
    await end_manual_command(interaction, "boss_toi")


@tree.command(
    name="liendaithu2",
    description="Bắt đầu/kết thúc điểm danh Liên Đài Thứ 2",
    guild=GUILD_SCOPE,
)
async def liendaithu2(interaction: discord.Interaction) -> None:
    await toggle_special_activity_command(interaction, "lien_dai_thu2")


@tree.command(
    name="liendaithu5",
    description="Bắt đầu/kết thúc điểm danh Liên Đài Thứ 5",
    guild=GUILD_SCOPE,
)
async def liendaithu5(interaction: discord.Interaction) -> None:
    await toggle_special_activity_command(interaction, "lien_dai_thu5")


@tree.command(
    name="ctc",
    description="Bắt đầu/kết thúc điểm danh CTC Thứ 7",
    guild=GUILD_SCOPE,
)
async def ctc(interaction: discord.Interaction) -> None:
    await toggle_special_activity_command(interaction, "ctc")


@tree.command(name="trangthai", description="Xem ca đang chạy và người trong voice", guild=GUILD_SCOPE)
async def trangthai(interaction: discord.Interaction) -> None:
    channel = await resolve_voice_channel()
    members = await voice_members(channel)
    active = await DB.active_runs()
    lines = ["**Trạng thái điểm danh**"]
    if active:
        for run in active:
            started = datetime.fromisoformat(run["started_at_utc"]).astimezone(CONFIG.tz)
            mode = "linh động" if run["mode"] == "manual" else "cố định"
            lines.append(f"🟢 {run['shift_label']} ({mode}) — bắt đầu {started:%H:%M:%S}")
    else:
        lines.append("⚪ Hiện không có ca nào đang chạy; bot không tính thời gian ngoài ca.")
    if members:
        lines.append("**Người đang trong voice:**")
        for member in members:
            stt = extract_stt(member.display_name)
            lines.append(f"• {member.display_name} — STT {stt if stt else 'không đọc được'}")
    else:
        lines.append("Hiện không có thành viên nào trong voice.")
    await interaction.response.send_message("\n".join(lines), ephemeral=True)


@tree.command(name="xuatexcel", description="Gửi file Excel hiện tại", guild=GUILD_SCOPE)
async def xuatexcel(interaction: discord.Interaction) -> None:
    if not can_control(interaction):
        await interaction.response.send_message("Bạn không có quyền dùng lệnh này.", ephemeral=True)
        return
    await interaction.response.send_message(
        "File Excel hiện tại:",
        file=discord.File(CONFIG.workbook_path, filename="Tuyet_Nguyet.xlsx"),
        ephemeral=True,
    )


@tree.command(name="capnhatexcel", description="Tải Excel mới lên và thay file bot đang dùng", guild=GUILD_SCOPE)
@app_commands.describe(file_excel="Chọn file .xlsx đã thêm/sửa thành viên")
async def capnhatexcel(interaction: discord.Interaction, file_excel: discord.Attachment) -> None:
    if not can_control(interaction):
        await interaction.response.send_message("Bạn không có quyền dùng lệnh này.", ephemeral=True)
        return

    if not file_excel.filename.lower().endswith(".xlsx"):
        await interaction.response.send_message(
            "Chỉ nhận file Excel định dạng `.xlsx`.", ephemeral=True
        )
        return

    if file_excel.size > 20 * 1024 * 1024:
        await interaction.response.send_message(
            "File quá lớn. Giới hạn của lệnh là 20 MB.", ephemeral=True
        )
        return

    active_runs = await DB.active_runs()
    if active_runs:
        labels = ", ".join(str(run["shift_label"]) for run in active_runs)
        await interaction.response.send_message(
            f"Không thể cập nhật Excel khi đang có ca chạy: {labels}. Hãy kết thúc ca trước.",
            ephemeral=True,
        )
        return

    await interaction.response.defer(ephemeral=True)
    upload_dir = CONFIG.workbook_path.parent / "uploads"
    upload_dir.mkdir(parents=True, exist_ok=True)
    temp_path = upload_dir / f"upload_{interaction.id}.xlsx"

    try:
        await file_excel.save(temp_path)
        backup, info = await WORKBOOK.install_uploaded_workbook(temp_path)
    except Exception as exc:
        log.exception("Cập nhật Excel thất bại")
        await interaction.followup.send(
            f"Không thể cập nhật Excel: `{exc}`\nFile cũ vẫn được giữ nguyên.",
            ephemeral=True,
        )
        return
    finally:
        try:
            temp_path.unlink(missing_ok=True)
        except OSError:
            pass

    backup_text = f" Bản cũ đã lưu tại `{backup.name}`." if backup.name else ""
    await interaction.followup.send(
        f"Đã cập nhật Excel thành công: **{info['members']} thành viên**, "
        f"STT lớn nhất **{info['max_stt']}**.{backup_text}",
        file=discord.File(CONFIG.workbook_path, filename="Tuyet_Nguyet.xlsx"),
        ephemeral=True,
    )


@tree.command(name="tuanmoi", description="Lưu bản cũ và xóa toàn bộ điểm danh E–AB", guild=GUILD_SCOPE)
@app_commands.describe(xac_nhan="Chọn True để xác nhận")
async def tuanmoi(interaction: discord.Interaction, xac_nhan: bool) -> None:
    if not can_control(interaction):
        await interaction.response.send_message("Bạn không có quyền dùng lệnh này.", ephemeral=True)
        return
    if not xac_nhan:
        await interaction.response.send_message("Chưa xóa. Chạy lại với `xac_nhan: True`.", ephemeral=True)
        return
    await interaction.response.defer(ephemeral=True)
    today = datetime.now(CONFIG.tz).date()
    monday = today - timedelta(days=today.weekday())
    archive = await WORKBOOK.new_week(monday)
    await DB.clear_week(monday)
    await interaction.followup.send(
        f"Đã tạo tuần mới. Bản cũ: `{archive.name}`.",
        file=discord.File(CONFIG.workbook_path, filename="Tuyet_Nguyet.xlsx"),
        ephemeral=True,
    )


def check_configuration() -> int:
    try:
        info = WORKBOOK.validate_sync()
    except Exception as exc:
        print(f"[LOI] {exc}")
        return 1
    print("[OK] Cấu hình hợp lệ")
    print(f"Server ID       : {CONFIG.guild_id}")
    print(f"Voice Channel ID: {CONFIG.voice_channel_id}")
    print(f"Text Channel ID : {CONFIG.report_channel_id}")
    print(f"Excel           : {CONFIG.workbook_path}")
    print(f"Sheet           : {info['sheet']}")
    print(f"Số thành viên   : {info['members']}")
    print(f"Ca linh động    : yêu cầu {CONFIG.manual_required_ratio:g} tổng thời gian")
    for shift in CONFIG.shifts:
        print(
            f"- {shift.label}: {shift.start:%H:%M}-{shift.end:%H:%M}, "
            f"tối thiểu {shift.minimum_minutes:g} phút nếu không có ca linh động"
        )
    print("Hoạt động riêng (gọi cùng lệnh để bắt đầu/kết thúc, yêu cầu 3/5):")
    for item in CONFIG.special_activities:
        print(
            f"- {item.label}: {weekday_text(date(2026, 7, 20) + timedelta(days=item.weekday))}, "
            f"cột {get_column_letter(item.column)}"
        )
    return 0


async def run_bot() -> None:
    await DB.initialize()
    await WORKBOOK.validate()
    load_dotenv(BASE_DIR / ".env")
    token = os.getenv("DISCORD_BOT_TOKEN", "").strip()
    if not token or token == "DAN_TOKEN_BOT_VAO_DAY":
        raise RuntimeError("Chưa có Bot Token trong file .env / Railway Variables.")
    try:
        await client.start(token)
    finally:
        try:
            await DB.heartbeat(datetime.now(UTC))
        finally:
            if client.is_closed() is False:
                await client.close()


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--check", action="store_true", help="Kiểm tra cấu hình và Excel")
    args = parser.parse_args()
    if args.check:
        return check_configuration()
    try:
        asyncio.run(run_bot())
    except KeyboardInterrupt:
        return 0
    except Exception as exc:
        log.exception("Bot dừng vì lỗi")
        print(f"\n[LOI] {exc}")
        return 1
    return 0


if __name__ == "__main__":
    sys.exit(main())
