from __future__ import annotations

import json
import shutil
import tempfile
import unicodedata
from pathlib import Path

from openpyxl import load_workbook

from attendance_core import extract_stt, manual_required_seconds


BASE = Path(__file__).resolve().parent


def normalize(value: object) -> str:
    text = str(value or "").strip().casefold()
    return "".join(
        char
        for char in unicodedata.normalize("NFD", text)
        if unicodedata.category(char) != "Mn"
    )


def rows_by_stt(ws, config: dict) -> dict[int, int]:
    result: dict[int, int] = {}
    marker = normalize(config["member_end_marker"])
    for row in range(config["first_member_row"], ws.max_row + 1):
        value = ws.cell(row=row, column=config["stt_column"]).value
        if normalize(value) == marker:
            break
        try:
            stt = int(float(value))
        except (TypeError, ValueError):
            continue
        if stt > 0:
            assert stt not in result, f"Trùng STT {stt}"
            result[stt] = row
    return result


def main() -> None:
    assert extract_stt("13 Babypi") == 13
    assert extract_stt("[13] Babypi") == 13
    assert extract_stt("Babypi 13") is None
    assert manual_required_seconds(1000) == 600

    config = json.loads((BASE / "config.json").read_text(encoding="utf-8"))
    assert config["guild_id"] == 1528460047904936016
    assert config["voice_channel_id"] == 1528460048404320394
    assert config["report_channel_id"] == 1528460048404320390
    assert config["sheet_name"] == "Trang tính2"

    expected_columns = {
        "boss_trua": [5, 9, 12, 15, 19, 22, 26],
        "van_tieu": [6, 10, 13, 16, 20, 23, 28],
        "boss_toi": [7, 11, 14, 17, 21, 24, 27],
        "lien_dai_thu2": [8, 8, 8, 8, 8, 8, 8],
        "lien_dai_thu5": [18, 18, 18, 18, 18, 18, 18],
        "ctc": [25, 25, 25, 25, 25, 25, 25],
    }
    assert config["attendance_columns"] == expected_columns

    template = BASE / "template" / "Tuyet_Nguyet.xlsx"
    workbook = load_workbook(template, data_only=False)
    worksheet = workbook[config["sheet_name"]]
    rows = rows_by_stt(worksheet, config)
    assert len(rows) == 72
    assert max(rows) == 72
    assert rows[1] == 3
    assert rows[72] == 74

    # Đúng cột ca trong form mới, kể cả Chủ nhật có thứ tự cột khác.
    assert worksheet.cell(2, 5).value == "BOSS 12H"
    assert "TIÊU" in str(worksheet.cell(2, 6).value).upper()
    assert "10H" in str(worksheet.cell(2, 7).value).upper()
    assert worksheet.cell(2, 26).value == "BOSS 12H"
    assert "22H" in str(worksheet.cell(2, 27).value).upper()
    assert "TIÊU" in str(worksheet.cell(2, 28).value).upper()

    with tempfile.TemporaryDirectory() as temp:
        copied = Path(temp) / "test.xlsx"
        shutil.copy2(template, copied)
        wb = load_workbook(copied)
        ws = wb[config["sheet_name"]]

        # Thêm thành viên mới ngay trên dòng tổng kết: bot phải tự nhận STT 73.
        marker_row = next(
            row
            for row in range(config["first_member_row"], ws.max_row + 1)
            if normalize(ws.cell(row, 1).value) == normalize(config["member_end_marker"])
        )
        ws.insert_rows(marker_row, 1)
        ws.cell(marker_row, 1).value = 73
        ws.cell(marker_row, 2).value = "Thành viên mới"
        new_rows = rows_by_stt(ws, config)
        assert len(new_rows) == 73
        assert new_rows[73] == marker_row

        # Ca Vận Tiêu linh động Chủ nhật phải thay đúng cột AB (28).
        sunday_van_tieu = config["attendance_columns"]["van_tieu"][6]
        ws.cell(new_rows[73], sunday_van_tieu).value = 1
        assert ws.cell(new_rows[73], 28).value == 1

        # Tạo tuần mới phải xóa toàn bộ vùng hoạt động E:AB, gồm Liên Đài và CTC.
        for col in range(5, 29):
            ws.cell(new_rows[1], col).value = 1
        attendance_cols = sorted({c for cols in expected_columns.values() for c in cols})
        assert attendance_cols == list(range(5, 29))
        for row in new_rows.values():
            for col in attendance_cols:
                ws.cell(row, col).value = None
        assert all(ws.cell(new_rows[1], col).value is None for col in range(5, 29))
        # Không xóa công thức tổng AC/AD.
        assert ws.cell(new_rows[1], 29).value is not None
        assert ws.cell(new_rows[1], 30).value is not None

        wb.save(copied)

    print("[OK] Form Excel, xóa tuần E:AB, Liên Đài/CTC và thêm thành viên đều hợp lệ.")


if __name__ == "__main__":
    main()
